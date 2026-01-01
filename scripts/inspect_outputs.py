import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

p=Path('outputs/no_trocr')
files=sorted([f for f in p.glob('*.xlsx')])
print('found',len(files),'xlsx files in',p)
if not files:
    print('no xlsx files found')
    raise SystemExit(0)

f=files[0]
print('first file:',f)

# Try reading with pandas first
try:
    df=pd.read_excel(f,header=None)
    if df.empty:
        print('\nPandas read returned empty DataFrame')
    else:
        print('\nFirst 10 rows (pandas):')
        print(df.head(10).to_string(index=False))
except Exception as e:
    print('pandas read error:',e)

# Inspect with openpyxl to see sheets and cell values
try:
    wb=load_workbook(f, read_only=True, data_only=True)
    print('\nWorksheets:', wb.sheetnames)
    for ws_name in wb.sheetnames:
        ws=wb[ws_name]
        print(f"\nSheet: {ws_name}")
        rows = ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=10, values_only=True)
        printed=False
        for r in rows:
            if any(cell is not None and str(cell).strip()!='' for cell in r):
                print('Row:', r)
                printed=True
        if not printed:
            print('  (no non-empty cells in first 10x10)')
    wb.close()
except Exception as e:
    print('openpyxl read error:',e)
